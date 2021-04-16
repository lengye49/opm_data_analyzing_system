import openpyxl

file_path = '../database/users.xlsx'
wb = openpyxl.load_workbook(file_path)
sh = wb['users']

id_list = []
type_list = []

for row in sh['A']:
    if row.value == 'Id':
        continue
    id_list.append(row.value)

for column in sh[1:1]:
    if column.value == 'Id':
        continue
    type_list.append(column.value)


def read_data(user_id, type_name):
    if user_id in id_list:
        _row = id_list.index(user_id) + 2
        _col = type_list.index(type_name) + 2
        return sh.cell(row=_row, column=_col).value
    else:
        return None


def save_data(user_id, type_name=None, value=None):
    if user_id not in id_list:
        id_list.append(user_id)
        _row = id_list.index(user_id) + 2
        _col = 1
        _ = sh.cell(row=_row, column=_col, value=user_id)

    if type_name is None:
        return
    else:
        _row = id_list.index(user_id) + 2
        _col = type_list.index(type_name) + 2
        _ = sh.cell(row=_row, column=_col, value=value)

    wb.save(file_path)


# print(read_data(10000004, 'Name'))
# save_data(10000008, 'Name', 'Test')
# print(read_data(10000005, 'Name'))
