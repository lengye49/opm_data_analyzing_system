"""
这个脚本用于从单个英雄信息表中读取指定信息
这个脚本跟save_hero_info.py 脚本对应使用

英雄信息表应该以英雄的id作为文件名
每个Sheet页命名有固定要求
"""

import openpyxl
import os

path = ''

HERO_CAMP = ['武装', '超能', '科技', '格斗', '全能']
HERO_ROLE = ['无畏', '敏捷', '战术']
HERO_JOB = ['重装', '近卫', '支援', '特种', '火力压制']


# 获取对应路径的Excel文件
def get_wb(hero_id):
    global path
    path = 'heroes/' + str(hero_id) + '.xlsx'
    if os.path.exists(path):
        return openpyxl.load_workbook(path)
    else:
        print('找不到英雄' + str(hero_id) + '的excel文件！')


# 获取Sheet页
def get_ws(wb, _name):
    if _name in wb.sheetnames:
        return wb[_name]
    else:
        print('找不到' + _name + '的Sheet！')


def load_hero_basic_info(hero_id):
    wb = get_wb(hero_id)
    ws = get_ws(wb, 'basic')

    hero_id = ws['B1'].value
    hero_name = ws['B2'].value
    hero_camp = HERO_CAMP.index(ws['B3'].value) + 1
    hero_profession = HERO_ROLE.index(ws['B4'].value) + 1
    hero_job = HERO_JOB.index(ws['B5'].value) + 30001
    hp_init = ws['B6'].value
    atk_init = ws['B7'].value
    def_init = ws['B8'].value
    crit_init = 500

    return hero_id, hero_name, hero_camp, hero_profession, hero_job, hp_init, atk_init, def_init, crit_init


def load_hero_quality_growth(hero_id):
    wb = get_wb(hero_id)
    ws = get_ws(wb, 'quality')

    hp_per = []
    atk_per = []
    def_per = []
    hp_v = []
    atk_v = []
    def_v = []

    hp_pve = []
    atk_pve = []
    def_pve = []

    hp_pve_aura = []
    atk_pve_aura = []
    def_pve_aura = []

    hp_pve_type_aura = []
    atk_pve_type_aura = []
    def_pve_type_aura = []

    hp_pvp = []
    atk_pvp = []
    def_pvp = []

    hp_pvp_aura = []
    atk_pvp_aura = []
    def_pvp_aura = []

    hp_pvp_type_aura = []
    atk_pvp_type_aura = []
    def_pvp_type_aura = []

    for i in range(0, 16):
        hp_per.append(ws['B' + str(i + 2)].value)
        atk_per.append(ws['C' + str(i + 2)].value)
        def_per.append(ws['D' + str(i + 2)].value)
        hp_v.append(ws['E' + str(i + 2)].value)
        atk_v.append(ws['F' + str(i + 2)].value)
        def_v.append(ws['G' + str(i + 2)].value)

        hp_pve.append(ws['N' + str(i + 2)].value)
        atk_pve.append(ws['O' + str(i + 2)].value)
        def_pve.append(ws['P' + str(i + 2)].value)

        hp_pve_aura.append(ws['Q' + str(i + 2)].value)
        atk_pve_aura.append(ws['R' + str(i + 2)].value)
        def_pve_aura.append(ws['S' + str(i + 2)].value)

        hp_pve_type_aura.append(ws['T' + str(i + 2)].value)
        atk_pve_type_aura.append(ws['U' + str(i + 2)].value)
        def_pve_type_aura.append(ws['V' + str(i + 2)].value)

        hp_pvp.append(ws['W' + str(i + 2)].value)
        atk_pvp.append(ws['X' + str(i + 2)].value)
        def_pvp.append(ws['Y' + str(i + 2)].value)

        hp_pvp_aura.append(ws['Z' + str(i + 2)].value)
        atk_pvp_aura.append(ws['AA' + str(i + 2)].value)
        def_pvp_aura.append(ws['AB' + str(i + 2)].value)

        hp_pvp_type_aura.append(ws['AC' + str(i + 2)].value)
        atk_pvp_type_aura.append(ws['AD' + str(i + 2)].value)
        def_pvp_type_aura.append(ws['AE' + str(i + 2)].value)

    return hp_per, atk_per, def_per, hp_v, atk_v, def_v, hp_pve, atk_pve, def_pve, \
           hp_pve_aura, atk_pve_aura, def_pve_aura, hp_pve_type_aura, atk_pve_type_aura, def_pve_type_aura, \
           hp_pvp, atk_pvp, def_pvp, hp_pvp_aura, atk_pvp_aura, def_pvp_aura, \
           hp_pvp_type_aura, atk_pvp_type_aura, def_pvp_type_aura


def load_hero_grade_growth(hero_id):
    wb = get_wb(hero_id)
    ws = get_ws(wb, 'grade')

    hp_v = []
    atk_v = []
    def_v = []
    for i in range(0, 20):
        hp_v.append(ws['B' + str(i + 2)].value)
        atk_v.append(ws['C' + str(i + 2)].value)
        def_v.append(ws['D' + str(i + 2)].value)

    return hp_v, atk_v, def_v


def load_hero_mechanical_power(hero_id):
    wb = get_wb(hero_id)
    ws = get_ws(wb, 'mechanical')

    hp_pve = []
    atk_pve = []
    def_pve = []
    hp_pvp = []
    atk_pvp = []
    def_pvp = []

    for i in range(0, 30):
        hp_pve.append(ws['B' + str(i + 2)].value)
        atk_pve.append(ws['C' + str(i + 2)].value)
        def_pve.append(ws['D' + str(i + 2)].value)
        hp_pvp.append(ws['E' + str(i + 2)].value)
        atk_pvp.append(ws['F' + str(i + 2)].value)
        def_pvp.append(ws['G' + str(i + 2)].value)

    return hp_pve, atk_pve, def_pve, hp_pvp, atk_pvp, def_pvp


def load_hero_talent(hero_id, talent_lv):
    wb = get_wb(hero_id)
    ws = get_ws(wb, 'talent')

    hp_per = []
    atk_per = []
    def_per = []
    crit = []
    crit_res = []
    precise = []
    parry = []
    dmg_res = []

    for i in range(0, talent_lv):
        hp_per.append(ws['B' + str(i + 2)].value )
        atk_per.append(ws['C' + str(i + 2)].value)
        def_per.append(ws['D' + str(i + 2)].value)
        crit.append(ws['E' + str(i + 2)].value)
        crit_res.append(ws['F' + str(i + 2)].value)
        precise.append(ws['G' + str(i + 2)].value)
        parry.append(ws['H' + str(i + 2)].value)
        dmg_res.append(ws['I' + str(i + 2)].value)

    return hp_per, atk_per, def_per, crit, crit_res, precise, parry, dmg_res


def load_hero_limiter(hero_id):
    wb = get_wb(hero_id)
    ws = get_ws(wb, 'limiter')

    hp_pve_v = []
    atk_pve_v = []
    def_pve_v = []
    hp_pve_per = []
    atk_pve_per = []
    def_pve_per = []

    hp_pve_aura = []
    atk_pve_aura = []
    def_pve_aura = []
    hp_pve_aura_per = []
    atk_pve_aura_per = []
    def_pve_aura_per = []

    hp_pve_type_aura = []
    atk_pve_type_aura = []
    def_pve_type_aura = []
    hp_pve_type_aura_per = []
    atk_pve_type_aura_per = []
    def_pve_type_aura_per = []

    hp_pvp_v = []
    atk_pvp_v = []
    def_pvp_v = []
    hp_pvp_per = []
    atk_pvp_per = []
    def_pvp_per = []

    hp_pvp_aura = []
    atk_pvp_aura = []
    def_pvp_aura = []
    hp_pvp_aura_per = []
    atk_pvp_aura_per = []
    def_pvp_aura_per = []

    hp_pvp_type_aura = []
    atk_pvp_type_aura = []
    def_pvp_type_aura = []
    hp_pvp_type_aura_per = []
    atk_pvp_type_aura_per = []
    def_pvp_type_aura_per = []

    for i in range(0, 10):
        hp_pve_v.append(ws['B' + str(i + 2)].value)
        atk_pve_v.append(ws['C' + str(i + 2)].value)
        def_pve_v.append(ws['D' + str(i + 2)].value)
        hp_pve_per.append(ws['E' + str(i + 2)].value)
        atk_pve_per.append(ws['F' + str(i + 2)].value)
        def_pve_per.append(ws['G' + str(i + 2)].value)

        hp_pve_aura.append(ws['H' + str(i + 2)].value)
        atk_pve_aura.append(ws['I' + str(i + 2)].value)
        def_pve_aura.append(ws['J' + str(i + 2)].value)
        hp_pve_aura_per.append(ws['K' + str(i + 2)].value)
        atk_pve_aura_per.append(ws['L' + str(i + 2)].value)
        def_pve_aura_per.append(ws['M' + str(i + 2)].value)

        hp_pve_type_aura.append(ws['N' + str(i + 2)].value)
        atk_pve_type_aura.append(ws['O' + str(i + 2)].value)
        def_pve_type_aura.append(ws['P' + str(i + 2)].value)
        hp_pve_type_aura_per.append(ws['Q' + str(i + 2)].value)
        atk_pve_type_aura_per.append(ws['R' + str(i + 2)].value)
        def_pve_type_aura_per.append(ws['S' + str(i + 2)].value)

        hp_pvp_v.append(ws['T' + str(i + 2)].value)
        atk_pvp_v.append(ws['U' + str(i + 2)].value)
        def_pvp_v.append(ws['V' + str(i + 2)].value)
        hp_pvp_per.append(ws['W' + str(i + 2)].value)
        atk_pvp_per.append(ws['X' + str(i + 2)].value)
        def_pvp_per.append(ws['Y' + str(i + 2)].value)

        hp_pvp_aura.append(ws['Z' + str(i + 2)].value)
        atk_pvp_aura.append(ws['AA' + str(i + 2)].value)
        def_pvp_aura.append(ws['AB' + str(i + 2)].value)
        hp_pvp_aura_per.append(ws['AC' + str(i + 2)].value)
        atk_pvp_aura_per.append(ws['AD' + str(i + 2)].value)
        def_pvp_aura_per.append(ws['AE' + str(i + 2)].value)

        hp_pvp_type_aura.append(ws['AF' + str(i + 2)].value)
        atk_pvp_type_aura.append(ws['AG' + str(i + 2)].value)
        def_pvp_type_aura.append(ws['AH' + str(i + 2)].value)
        hp_pvp_type_aura_per.append(ws['AI' + str(i + 2)].value)
        atk_pvp_type_aura_per.append(ws['AJ' + str(i + 2)].value)
        def_pvp_type_aura_per.append(ws['AK' + str(i + 2)].value)

    return hp_pve_v, hp_pve_per, hp_pve_aura, hp_pve_aura_per, hp_pve_type_aura, hp_pve_type_aura_per, \
           hp_pvp_v, hp_pvp_per, hp_pvp_aura, hp_pvp_aura_per,hp_pvp_type_aura,hp_pvp_type_aura_per, atk_pve_v,\
           atk_pve_per, atk_pve_aura, atk_pve_aura_per,atk_pve_type_aura, atk_pve_type_aura_per, atk_pvp_v,\
           atk_pvp_per, atk_pvp_aura,atk_pvp_aura_per, atk_pvp_type_aura, atk_pvp_type_aura_per, def_pve_v,\
           def_pve_per,def_pve_aura, def_pve_aura_per, def_pve_type_aura, def_pve_type_aura_per,\
           def_pvp_v,def_pvp_per, def_pvp_aura, def_pvp_aura_per, def_pvp_type_aura, def_pvp_type_aura_per

