import openpyxl

hero_list = [1, 2, 8, 10, 11, 12, 13, 18, 19, 20, 21, 23, 24, 25, 26, 27, 31, 32, 33, 34, 35, 36, 38, 39, 40, 41, 42,
             43, 44, 45, 46, 49, 51, 60, 61, 62, 63, 83, 84, 85, 86, 87, 88, 89, 90, 92, 93, 94, 95, 96, 97, 98, 100,
             101, 102, 103, 104, 105, 107, 108, 109, 110, 111, 112, 113, 114, 115, 116, 117, 118]
id_info = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 20, 21, 22, 23, 24, 25]

wb_status = openpyxl.load_workbook('heroes/status.xlsx')
ws_status = wb_status['Sheet']

i = 1
for hero in hero_list:
    print('processing hero %d' % hero)
    _p = 'heroes/' + str(hero) + '.xlsx'
    wb = openpyxl.load_workbook(_p)
    ws = wb['status']

    for j in range(2, 24):
        ws_status['A' + str(i)] = 100000 + id_info[j - 2] * 1000 + hero

        ws_status['B' + str(i)] = ws['L' + str(j + 2)].value
        ws_status['C' + str(i)] = ws['M' + str(j + 2)].value
        ws_status['D' + str(i)] = ws['N' + str(j + 2)].value
        ws_status['E' + str(i)] = ws['O' + str(j + 2)].value
        ws_status['F' + str(i)] = ws['P' + str(j + 2)].value
        ws_status['G' + str(i)] = ws['Q' + str(j + 2)].value
        ws_status['H' + str(i)] = ws['R' + str(j + 2)].value
        ws_status['I' + str(i)] = ws['S' + str(j + 2)].value
        ws_status['J' + str(i)] = ws['T' + str(j + 2)].value
        ws_status['K' + str(i)] = ws['U' + str(j + 2)].value
        ws_status['L' + str(i)] = ws['V' + str(j + 2)].value
        ws_status['M' + str(i)] = ws['W' + str(j + 2)].value
        ws_status['N' + str(i)] = ws['X' + str(j + 2)].value
        ws_status['O' + str(i)] = ws['Y' + str(j + 2)].value
        ws_status['P' + str(i)] = ws['Z' + str(j + 2)].value
        ws_status['Q' + str(i)] = ws['AA' + str(j + 2)].value
        ws_status['R' + str(i)] = ws['AB' + str(j + 2)].value
        ws_status['S' + str(i)] = ws['AC' + str(j + 2)].value
        ws_status['T' + str(i)] = ws['AD' + str(j + 2)].value
        ws_status['U' + str(i)] = ws['AE' + str(j + 2)].value
        ws_status['V' + str(i)] = ws['AF' + str(j + 2)].value
        ws_status['W' + str(i)] = ws['AG' + str(j + 2)].value
        ws_status['X' + str(i)] = ws['AH' + str(j + 2)].value
        ws_status['Y' + str(i)] = ws['AI' + str(j + 2)].value
        ws_status['Z' + str(i)] = ws['AJ' + str(j + 2)].value
        ws_status['AA' + str(i)] = ws['AK' + str(j + 2)].value
        ws_status['AB' + str(i)] = ws['AL' + str(j + 2)].value
        ws_status['AC' + str(i)] = ws['AM' + str(j + 2)].value
        ws_status['AD' + str(i)] = ws['AN' + str(j + 2)].value
        i += 1

    wb.close()

wb_status.save('heroes/status.xlsx')
