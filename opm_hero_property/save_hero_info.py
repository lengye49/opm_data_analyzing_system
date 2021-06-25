import openpyxl
import os

HERO_CAMP = ['武装', '超能', '科技', '格斗', '全能']
HERO_ROLE = ['无畏', '敏捷', '战术']
HERO_JOB = ['重装', '近卫', '支援', '特种', '火力压制']

path = ''


def get_wb(hero_id):
    global path
    path = 'heroes/' + str(hero_id) + '.xlsx'
    if os.path.exists(path):
        return openpyxl.load_workbook(path)
    else:
        return openpyxl.Workbook()


def get_ws(wb, _name):
    if _name in wb.sheetnames:
        return wb[_name]
    else:
        return wb.create_sheet(_name)


def save_hero_basic_info(hero_id, hero_name, hero_camp, hero_profession, hero_job, hp_init, atk_init, def_init):
    wb = get_wb(hero_id)
    ws = get_ws(wb, 'basic')

    ws['A1'] = '英雄ID'
    ws['B1'] = hero_id

    ws['A2'] = '英雄名称'
    ws['B2'] = hero_name

    ws['A3'] = '阵营'
    ws['B3'] = HERO_CAMP[hero_camp - 1]

    ws['A4'] = '定位'
    ws['B4'] = HERO_ROLE[hero_profession - 1]

    ws['A5'] = '职阶'
    ws['B5'] = HERO_JOB[hero_job - 30001]

    ws['A6'] = '初始生命'
    ws['B6'] = hp_init

    ws['A7'] = '初始攻击'
    ws['B7'] = atk_init

    ws['A8'] = '初始防御'
    ws['B8'] = def_init

    ws['A9'] = '初始暴击'
    ws['B9'] = 500

    wb.save(path)


def save_hero_quality_growth(hero_id, hp_per, atk_per, def_per, hp_v, atk_v, def_v):
    wb = get_wb(hero_id)
    ws = get_ws(wb, 'quality')

    ws['A1'] = '品质ID'
    ws['B1'] = '生命成长'
    ws['C1'] = '攻击成长'
    ws['D1'] = '防御成长'
    ws['E1'] = '生命增加'
    ws['F1'] = '攻击增加'
    ws['G1'] = '防御增加'

    for i in range(0, 16):
        ws['A' + str(i + 2)] = i + 1
        ws['B' + str(i + 2)] = hp_per[i]
        ws['C' + str(i + 2)] = atk_per[i]
        ws['D' + str(i + 2)] = def_per[i]
        ws['E' + str(i + 2)] = hp_v[i]
        ws['F' + str(i + 2)] = atk_v[i]
        ws['G' + str(i + 2)] = def_v[i]

    wb.save(path)


def save_hero_quality_max_value(hero_id, hp_pve, atk_pve, def_pve, hp_pvp, atk_pvp, def_pvp,
                                hp_pve_aura, atk_pve_aura, def_pve_aura,
                                hp_pvp_aura, atk_pvp_aura, def_pvp_aura,
                                hp_pve_type_aura, atk_pve_type_aura, def_pve_type_aura,
                                hp_pvp_type_aura, atk_pvp_type_aura, def_pvp_type_aura, ):
    wb = get_wb(hero_id)
    ws = get_ws(wb, 'quality')

    ws['H1'] = 'PVE属性最大值'
    ws['I1'] = 'PVE光环最大值'
    ws['J1'] = 'PVE类型光环最大值'
    ws['K1'] = 'PVP属性最大值'
    ws['L1'] = 'PVP属性最大值'
    ws['M1'] = 'PVP属性最大值'

    ws['N1'] = 'PVE属性最大值_HP'
    ws['O1'] = 'PVE属性最大值_ATK'
    ws['P1'] = 'PVE属性最大值_DEF'

    ws['Q1'] = 'PVE光环最大值'
    ws['R1'] = 'PVE光环最大值'
    ws['S1'] = 'PVE光环最大值'

    ws['T1'] = 'PVE阵营光环最大值'
    ws['U1'] = 'PVE阵营光环最大值'
    ws['V1'] = 'PVE阵营光环最大值'

    ws['W1'] = 'PVP属性最大值_HP'
    ws['X1'] = 'PVP属性最大值_ATK'
    ws['Y1'] = 'PVP属性最大值_DEF'

    ws['Z1'] = 'PVP光环最大值'
    ws['AA1'] = 'PVP光环最大值'
    ws['AB1'] = 'PVP光环最大值'

    ws['AC1'] = 'PVP阵营光环最大值'
    ws['AD1'] = 'PVP阵营光环最大值'
    ws['AE1'] = 'PVP阵营光环最大值'

    for i in range(0, 16):
        s1 = '21,' + str(int(hp_pve[i])) + ';31,' + str(int(atk_pve[i])) + ';41,' + str(int(def_pve[i]))
        s2 = '21,' + str(int(hp_pve_aura[i])) + ';31,' + str(int(atk_pve_aura[i])) + ';41,' + str(
            int(def_pve_aura[i]))
        s3 = '21,' + str(int(hp_pve_type_aura[i])) + ';31,' + str(int(atk_pve_type_aura[i])) + ';41,' + str(
            int(def_pve_type_aura[i]))
        s4 = '21,' + str(int(hp_pvp[i])) + ';31,' + str(int(atk_pvp[i])) + ';41,' + str(int(def_pvp[i]))
        s5 = '21,' + str(int(hp_pvp_aura[i])) + ';31,' + str(int(atk_pvp_aura[i])) + ';41,' + str(
            int(def_pvp_aura[i]))
        s6 = '21,' + str(int(hp_pvp_type_aura[i])) + ';31,' + str(int(atk_pvp_type_aura[i])) + ';41,' + str(
            int(def_pvp_type_aura[i]))

        ws['H' + str(i + 2)] = s1
        ws['I' + str(i + 2)] = s2
        ws['J' + str(i + 2)] = s3
        ws['K' + str(i + 2)] = s4
        ws['L' + str(i + 2)] = s5
        ws['M' + str(i + 2)] = s6

        ws['N' + str(i + 2)] = hp_pve[i]
        ws['O' + str(i + 2)] = atk_pve[i]
        ws['P' + str(i + 2)] = def_pve[i]

        ws['Q' + str(i + 2)] = hp_pve_aura[i]
        ws['R' + str(i + 2)] = atk_pve_aura[i]
        ws['S' + str(i + 2)] = def_pve_aura[i]

        ws['T' + str(i + 2)] = hp_pve_type_aura[i]
        ws['U' + str(i + 2)] = atk_pve_type_aura[i]
        ws['V' + str(i + 2)] = def_pve_type_aura[i]

        ws['W' + str(i + 2)] = hp_pvp[i]
        ws['X' + str(i + 2)] = atk_pvp[i]
        ws['Y' + str(i + 2)] = def_pvp[i]

        ws['Z' + str(i + 2)] = hp_pvp_aura[i]
        ws['AA' + str(i + 2)] = atk_pvp_aura[i]
        ws['AB' + str(i + 2)] = def_pvp_aura[i]

        ws['AC' + str(i + 2)] = hp_pvp_type_aura[i]
        ws['AD' + str(i + 2)] = atk_pvp_type_aura[i]
        ws['AE' + str(i + 2)] = def_pvp_type_aura[i]

    wb.save(path)


def save_hero_grade_growth(hero_id, hp_v, atk_v, def_v):
    wb = get_wb(hero_id)
    ws = get_ws(wb, 'grade')

    ws['A1'] = '品阶ID'
    ws['B1'] = '生命增加'
    ws['C1'] = '攻击增加'
    ws['D1'] = '防御增加'

    for i in range(0, 20):
        ws['A' + str(i + 2)] = i + 1
        ws['B' + str(i + 2)] = hp_v[i]
        ws['C' + str(i + 2)] = atk_v[i]
        ws['D' + str(i + 2)] = def_v[i]

    wb.save(path)


def save_hero_mechanical_power(hero_id, hp_pve, atk_pve, def_pve, hp_pvp, atk_pvp, def_pvp):
    wb = get_wb(hero_id)
    ws = get_ws(wb, 'mechanical')
    ws['A1'] = '等级'
    ws['B1'] = '生命_pve'
    ws['C1'] = '攻击_pve'
    ws['D1'] = '防御_pve'
    ws['E1'] = '生命_pvp'
    ws['F1'] = '攻击_pvp'
    ws['G1'] = '防御_pvp'
    ws['H1'] = 'pve输出'
    ws['I1'] = 'pvp输出'

    for i in range(0, 30):
        ws['A' + str(i + 2)] = i + 1
        ws['B' + str(i + 2)] = hp_pve[i]
        ws['C' + str(i + 2)] = atk_pve[i]
        ws['D' + str(i + 2)] = def_pve[i]
        ws['E' + str(i + 2)] = hp_pvp[i]
        ws['F' + str(i + 2)] = atk_pvp[i]
        ws['G' + str(i + 2)] = def_pvp[i]
        s1 = '21,' + str(hp_pve[i]) + ';31,' + str(atk_pve[i]) + ';41,' + str(def_pve[i])
        s2 = '21,' + str(hp_pvp[i]) + ';31,' + str(atk_pvp[i]) + ';41,' + str(def_pvp[i])
        ws['H' + str(i + 2)] = s1
        ws['I' + str(i + 2)] = s2

    wb.save(path)


def save_hero_talent_growth(hero_id, talent_lv, hp_per, atk_per, def_per, crit, crit_res, precise, parry, dmg_res):
    wb = get_wb(hero_id)
    ws = get_ws(wb, 'talent')

    ws['A1'] = '天赋等级'
    ws['B1'] = '生命%'
    ws['C1'] = '攻击%'
    ws['D1'] = '防御%'
    ws['E1'] = '暴击'
    ws['F1'] = '暴击抵抗'
    ws['G1'] = '精准'
    ws['H1'] = '格挡'
    ws['I1'] = '伤害减免'
    ws['J1'] = '导出数据'

    for i in range(0, talent_lv):
        ws['A' + str(i + 2)] = i
        ws['B' + str(i + 2)] = round(hp_per[i] * 100)
        ws['C' + str(i + 2)] = round(atk_per[i] * 100)
        ws['D' + str(i + 2)] = round(def_per[i] * 100)
        ws['E' + str(i + 2)] = crit[i]
        ws['F' + str(i + 2)] = crit_res[i]
        ws['G' + str(i + 2)] = precise[i]
        ws['H' + str(i + 2)] = parry[i]
        ws['I' + str(i + 2)] = dmg_res[i]
        s = ';'
        if hp_per[i] > 0:
            s += '22,' + str(int(round(hp_per[i] * 100))) + ';'
        if atk_per[i] > 0:
            s += '32,' + str(int(round(atk_per[i] * 100))) + ';'
        if def_per[i] > 0:
            s += '42,' + str(int(round(def_per[i] * 100))) + ';'
        if crit[i] > 0:
            s += '70,' + str(int(crit[i])) + ';'
        if crit_res[i] > 0:
            s += '60,' + str(int(crit_res[i])) + ';'
        if precise[i] > 0:
            s += '150,' + str(int(precise[i])) + ';'
        if parry[i] > 0:
            s += '140,' + str(int(parry[i])) + ';'
        if dmg_res[i] > 0:
            s += '50,' + str(int(dmg_res[i])) + ';'
        ws['J' + str(i + 2)] = s[1:-1:]

    wb.save(path)


def save_hero_limiter_growth(hero_id, hp_pve_v, hp_pve_per, hp_pve_aura, hp_pve_aura_per, hp_pve_type_aura,
                             hp_pve_type_aura_per, hp_pvp_v, hp_pvp_per, hp_pvp_aura, hp_pvp_aura_per, hp_pvp_type_aura,
                             hp_pvp_type_aura_per, atk_pve_v, atk_pve_per, atk_pve_aura, atk_pve_aura_per,
                             atk_pve_type_aura, atk_pve_type_aura_per, atk_pvp_v, atk_pvp_per, atk_pvp_aura,
                             atk_pvp_aura_per, atk_pvp_type_aura, atk_pvp_type_aura_per, def_pve_v, def_pve_per,
                             def_pve_aura, def_pve_aura_per, def_pve_type_aura, def_pve_type_aura_per, def_pvp_v,
                             def_pvp_per, def_pvp_aura, def_pvp_aura_per, def_pvp_type_aura, def_pvp_type_aura_per):
    wb = get_wb(hero_id)
    ws = get_ws(wb, 'limiter')

    ws['A1'] = '限制器等级'
    ws['B1'] = 'pve属性生命'
    ws['C1'] = 'pve属性攻击'
    ws['D1'] = 'pve属性防御'
    ws['E1'] = 'pve属性生命%'
    ws['F1'] = 'pve属性攻击%'
    ws['G1'] = 'pve属性防御%'

    ws['H1'] = 'pve光环生命'
    ws['I1'] = 'pve光环攻击'
    ws['J1'] = 'pve光环防御'
    ws['K1'] = 'pve光环生命%'
    ws['L1'] = 'pve光环攻击%'
    ws['M1'] = 'pve光环防御%'

    ws['N1'] = 'pve阵营光环生命'
    ws['O1'] = 'pve阵营光环攻击'
    ws['P1'] = 'pve阵营光环防御'
    ws['Q1'] = 'pve阵营光环生命%'
    ws['R1'] = 'pve阵营光环攻击%'
    ws['S1'] = 'pve阵营光环防御%'

    ws['T1'] = 'pvp属性生命'
    ws['U1'] = 'pvp属性攻击'
    ws['V1'] = 'pvp属性防御'
    ws['W1'] = 'pvp属性生命%'
    ws['X1'] = 'pvp属性攻击%'
    ws['Y1'] = 'pvp属性防御%'

    ws['Z1'] = 'pvp光环生命'
    ws['AA1'] = 'pvp光环攻击'
    ws['AB1'] = 'pvp光环防御'
    ws['AC1'] = 'pvp光环生命%'
    ws['AD1'] = 'pvp光环攻击%'
    ws['AE1'] = 'pvp光环防御%'

    ws['AF1'] = 'pvp阵营光环生命'
    ws['AG1'] = 'pvp阵营光环攻击'
    ws['AH1'] = 'pvp阵营光环防御'
    ws['AI1'] = 'pvp阵营光环生命%'
    ws['AJ1'] = 'pvp阵营光环攻击%'
    ws['AK1'] = 'pvp阵营光环防御%'

    ws['AL1'] = 'pve属性输出'
    ws['AM1'] = 'pve光环输出'
    ws['AN1'] = 'pve阵营光环输出'
    ws['AO1'] = 'pvp属性输出'
    ws['AP1'] = 'pvp光环输出'
    ws['AQ1'] = 'pvp阵营光环输出'

    for i in range(0, 10):
        ws['A' + str(i + 2)] = i + 1
        ws['B' + str(i + 2)] = hp_pve_v[i]
        ws['C' + str(i + 2)] = atk_pve_v[i]
        ws['D' + str(i + 2)] = def_pve_v[i]
        ws['E' + str(i + 2)] = hp_pve_per[i]
        ws['F' + str(i + 2)] = atk_pve_per[i]
        ws['G' + str(i + 2)] = def_pve_per[i]
        s1 = '21,' + str(hp_pve_v[i]) + ';31,' + str(atk_pve_v[i]) + ';41,' + str(def_pve_v[i]) + ';22,' + str(
            hp_pve_per[i]) + ';32,' + str(atk_pve_per[i]) + ';42,' + str(def_pve_per[i])

        ws['H' + str(i + 2)] = hp_pve_aura[i]
        ws['I' + str(i + 2)] = atk_pve_aura[i]
        ws['J' + str(i + 2)] = def_pve_aura[i]
        ws['K' + str(i + 2)] = hp_pve_aura_per[i]
        ws['L' + str(i + 2)] = atk_pve_aura_per[i]
        ws['M' + str(i + 2)] = def_pve_aura_per[i]
        s2 = '21,' + str(hp_pve_aura[i]) + ';31,' + str(atk_pve_aura[i]) + ';41,' + str(def_pve_aura[i]) + ';22,' + str(
            hp_pve_aura_per[i]) + ';32,' + str(atk_pve_aura_per[i]) + ';42,' + str(def_pve_aura_per[i])

        ws['N' + str(i + 2)] = hp_pve_type_aura[i]
        ws['O' + str(i + 2)] = atk_pve_type_aura[i]
        ws['P' + str(i + 2)] = def_pve_type_aura[i]
        ws['Q' + str(i + 2)] = hp_pve_type_aura_per[i]
        ws['R' + str(i + 2)] = atk_pve_type_aura_per[i]
        ws['S' + str(i + 2)] = def_pve_type_aura_per[i]
        s3 = '21,' + str(hp_pve_type_aura[i]) + ';31,' + str(atk_pve_type_aura[i]) + ';41,' + str(
            def_pve_type_aura[i]) + ';22,' + str(
            hp_pve_type_aura_per[i]) + ';32,' + str(atk_pve_type_aura_per[i]) + ';42,' + str(def_pve_type_aura_per[i])

        ws['T' + str(i + 2)] = hp_pvp_v[i]
        ws['U' + str(i + 2)] = atk_pvp_v[i]
        ws['V' + str(i + 2)] = def_pvp_v[i]
        ws['W' + str(i + 2)] = hp_pvp_per[i]
        ws['X' + str(i + 2)] = atk_pvp_per[i]
        ws['Y' + str(i + 2)] = def_pvp_per[i]
        s4 = '21,' + str(hp_pvp_v[i]) + ';31,' + str(atk_pvp_v[i]) + ';41,' + str(def_pvp_v[i]) + ';22,' + str(
            hp_pvp_per[i]) + ';32,' + str(atk_pvp_per[i]) + ';42,' + str(def_pvp_per[i])

        ws['Z' + str(i + 2)] = hp_pvp_aura[i]
        ws['AA' + str(i + 2)] = atk_pvp_aura[i]
        ws['AB' + str(i + 2)] = def_pvp_aura[i]
        ws['AC' + str(i + 2)] = hp_pvp_aura_per[i]
        ws['AD' + str(i + 2)] = atk_pvp_aura_per[i]
        ws['AE' + str(i + 2)] = def_pvp_aura_per[i]
        s5 = '21,' + str(hp_pvp_aura[i]) + ';31,' + str(atk_pvp_aura[i]) + ';41,' + str(def_pvp_aura[i]) + ';22,' + str(
            hp_pvp_aura_per[i]) + ';32,' + str(atk_pvp_aura_per[i]) + ';42,' + str(def_pvp_aura_per[i])

        ws['AF' + str(i + 2)] = hp_pvp_type_aura[i]
        ws['AG' + str(i + 2)] = atk_pvp_type_aura[i]
        ws['AH' + str(i + 2)] = def_pvp_type_aura[i]
        ws['AI' + str(i + 2)] = hp_pvp_type_aura_per[i]
        ws['AJ' + str(i + 2)] = atk_pvp_type_aura_per[i]
        ws['AK' + str(i + 2)] = def_pvp_type_aura_per[i]
        s6 = '21,' + str(hp_pvp_type_aura[i]) + ';31,' + str(atk_pvp_type_aura[i]) + ';41,' + str(
            def_pvp_type_aura[i]) + ';22,' + str(hp_pvp_type_aura_per[i]) + ';32,' + str(
            atk_pvp_type_aura_per[i]) + ';42,' + str(def_pvp_type_aura_per[i])

        ws['AL' + str(i + 2)] = s1
        ws['AM' + str(i + 2)] = s2
        ws['AN' + str(i + 2)] = s3
        ws['AO' + str(i + 2)] = s4
        ws['AP' + str(i + 2)] = s5
        ws['AQ' + str(i + 2)] = s6

    wb.save(path)


def get_skills(_level, _talent, _job, _limiter, _limiter_on):
    s = ''
    if _level > 220:
        s = '5,3,3,3'
    elif _level > 200:
        s = '4,3,3,3'
    elif _level > 180:
        s = '4,3,3,2'
    elif _level > 160:
        s = '4,3,2,2'
    elif _level > 140:
        s = '4,2,2,2'
    elif _level > 120:
        s = '3,2,2,2'
    elif _level > 100:
        s = '3,2,2,1'
    elif _level > 80:
        s = '3,2,1,1'
    elif _level > 60:
        s = '3,1,1,1'
    elif _level > 40:
        s = '2,1,1,1'
    elif _level > 20:
        s = '2,1,1,0'
    elif _level > 10:
        s = '1,1,1,0'
    else:
        s = '1,1,0,0'

    if _talent > 29:
        s += ',4'
    elif _talent > 19:
        s += ',3'
    elif _talent > 9:
        s += ',2'
    elif _talent >= 0:
        s += ',1'
    else:
        s += ',0'

    if _job > 74:
        s += ',5,5,5'
    elif _job > 69:
        s += ',5,5,4'
    elif _job > 64:
        s += ',5,4,4'
    elif _job > 59:
        s += ',4,4,4'
    elif _job > 54:
        s += ',4,4,3'
    elif _job > 49:
        s += ',4,3,3'
    elif _job > 44:
        s += ',3,3,3'
    elif _job > 39:
        s += ',3,3,2'
    elif _job > 34:
        s += ',3,2,2'
    elif _job > 29:
        s += ',2,2,2'
    elif _job > 24:
        s += ',2,2,1'
    elif _job > 19:
        s += ',2,2,0'
    elif _job > 14:
        s += ',2,1,0'
    elif _job > 9:
        s += ',2,0,0'
    elif _job > 4:
        s += ',1,0,0'
    else:
        s += ',0,0,0'

    if _limiter_on:
        s += ',' + str(_limiter)

    return s


def save_hero_status(hero_id, levels, df, hp, atk, de, crit, crit_res, crit_dmg, precise, parry, dmg_res, aura,
                     type_aura, _p_base, _p_equip, _p_talent, _p_academy, _p_job, _p_mechanical, _p_limiter,
                     _p_collection, _p_total, limiter_on, job_contact, pvp_addition):
    wb = get_wb(hero_id)
    ws = get_ws(wb, 'status')

    ws['A1'] = '战斗类型'
    ws['B1'] = '品质'
    ws['C1'] = 'pve等级'
    ws['D1'] = '等级强化'
    ws['E1'] = '装备品质%'
    ws['F1'] = '装备强化%'
    ws['G1'] = '天赋'
    ws['H1'] = '研究所核心'
    ws['I1'] = '职阶'
    ws['J1'] = '限制器'
    ws['K1'] = '机械核心'

    ws['L1'] = 'HeroId'
    ws['M1'] = 'Type'
    ws['N1'] = 'Quality'
    ws['O1'] = 'Level'
    ws['P1'] = 'EnhancePeriod'
    ws['Q1'] = 'Equips'
    ws['R1'] = 'TalentLevel'
    ws['S1'] = 'Skills'

    ws['T1'] = 'hp'
    ws['U1'] = 'atk'
    ws['V1'] = 'def'
    ws['W1'] = 'crit'
    ws['X1'] = 'crit_res'
    ws['Y1'] = 'crit_dmg'
    ws['Z1'] = 'precise'
    ws['AA1'] = 'parry'
    ws['AB1'] = 'dmg_res'
    ws['AC1'] = 'fury'

    ws['AD1'] = 'show_level'
    ws['AE1'] = 'chase'
    ws['AF1'] = 'academy'
    ws['AG1'] = 'job'
    ws['AH1'] = 'limiter'
    ws['AI1'] = 'mechanical'
    ws['AJ1'] = 'aura'
    ws['AK1'] = 'type_aura'
    ws['AL1'] = 'job_contact'
    ws['AM1'] = 'attr_addition'
    ws['AN1'] = 'pvp_addition'

    ws['AO1'] = '基础战力'
    ws['AP1'] = '装备战力'
    ws['AQ1'] = '天赋战力'
    ws['AR1'] = '研究所战力'
    ws['AS1'] = '职阶战力'
    ws['AT1'] = '机械核心战力'
    ws['AU1'] = '限制器战力'
    ws['AV1'] = '收集战力'
    ws['AW1'] = '总战力'

    ll = df.index.values
    for i in range(0, len(ll)):
        k = ll[i]
        ws['A' + str(i + 2)] = df.loc[k, '_type']
        ws['B' + str(i + 2)] = df.loc[k, '_quality']
        ws['C' + str(i + 2)] = df.loc[k, '_pve_level']
        ws['D' + str(i + 2)] = df.loc[k, '_lv_enhance']
        ws['E' + str(i + 2)] = df.loc[k, '_e_qua']
        ws['F' + str(i + 2)] = df.loc[k, '_e_enhance']
        ws['G' + str(i + 2)] = df.loc[k, '_talent']
        ws['H' + str(i + 2)] = df.loc[k, '_core']
        ws['I' + str(i + 2)] = df.loc[k, '_job']
        ws['J' + str(i + 2)] = df.loc[k, '_limiter']
        ws['K' + str(i + 2)] = df.loc[k, '_mechanical']

        ws['L' + str(i + 2)] = hero_id
        ws['M' + str(i + 2)] = 2
        ws['N' + str(i + 2)] = df.loc[k, '_quality']
        ws['O' + str(i + 2)] = levels[i]
        ws['P' + str(i + 2)] = 0
        ws['Q' + str(i + 2)] = ''
        ws['R' + str(i + 2)] = df.loc[k, '_talent']
        ws['S' + str(i + 2)] = get_skills(levels[i], df.loc[k, '_talent'], df.loc[k, '_job'], df.loc[k, '_limiter'], limiter_on)

        ws['T' + str(i + 2)] = hp[i]
        ws['U' + str(i + 2)] = atk[i]
        ws['V' + str(i + 2)] = de[i]
        ws['W' + str(i + 2)] = crit[i]
        ws['X' + str(i + 2)] = crit_res[i]
        ws['Y' + str(i + 2)] = crit_dmg[i]
        ws['Z' + str(i + 2)] = precise[i]
        ws['AA' + str(i + 2)] = parry[i]
        ws['AB' + str(i + 2)] = dmg_res[i]
        ws['AC' + str(i + 2)] = 0

        ws['AD' + str(i + 2)] = levels[i]
        ws['AE' + str(i + 2)] = 5000
        ws['AF' + str(i + 2)] = df.loc[k, '_core']
        ws['AG' + str(i + 2)] = df.loc[k, '_job']
        ws['AH' + str(i + 2)] = df.loc[k, '_limiter']
        ws['AI' + str(i + 2)] = df.loc[k, '_mechanical']
        ws['AJ' + str(i + 2)] = aura[i]
        ws['AK' + str(i + 2)] = type_aura[i]
        ws['AL' + str(i + 2)] = job_contact[i]
        ws['AM' + str(i + 2)] = ''
        ws['AN' + str(i + 2)] = pvp_addition[i]

        ws['AO' + str(i + 2)] = _p_base[i]
        ws['AP' + str(i + 2)] = _p_equip[i]
        ws['AQ' + str(i + 2)] = _p_talent[i]
        ws['AR' + str(i + 2)] = _p_academy[i]
        ws['AS' + str(i + 2)] = _p_job[i]
        ws['AT' + str(i + 2)] = _p_mechanical[i]
        ws['AU' + str(i + 2)] = _p_limiter[i]
        ws['AV' + str(i + 2)] = _p_collection[i]
        ws['AW' + str(i + 2)] = _p_total[i]

    wb.save(path)
