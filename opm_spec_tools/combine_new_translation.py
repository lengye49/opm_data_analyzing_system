import openpyxl
import openpyxl.styles
from openpyxl.styles import PatternFill


path = 'translation/【手游】一拳超人：英雄之路2.0_mpsen_v2.3.0_新增翻译 '
path_tr = 'languages.xlsx'
langs = ['英语', '俄语', '土语', '德语', '法语', '西语', '葡语']
fil = PatternFill('solid', fgColor='FFBB02')


print('加载文件中...\n')
wb = openpyxl.load_workbook(path_tr)
for lang in langs:
    ws_result = wb[lang]
    print('正在合并', lang, '...')
    try:
        wb_ori = openpyxl.load_workbook(path + lang + '.xlsx')
        ws_ori = wb_ori['Sheet1']
    except IOError:
        print('找不到文件：', path + lang + '.xlsx')
        continue

    # 读取翻译文件
    lang_dict = {}
    for i in range(2, ws_ori.max_row + 1):
        if ws_ori['A' + str(i)].value:
            lang_dict[ws_ori['A' + str(i)].value] = ws_ori['E' + str(i)].value
    wb_ori.close()

    # 替换文件
    j = 1
    for i in range(1, ws_result.max_row + 1):
        if ws_result['A' + str(i)].value is None:
            continue
        if ws_result['A' + str(i)].value in lang_dict:
            ws_result['B' + str(i)].value = lang_dict[ws_result['A' + str(i)].value]
            lang_dict.pop(ws_result['A' + str(i)].value)
            ws_result['A' + str(i)].fill = fil
            ws_result['B' + str(i)].fill = fil
        if int(ws_result['A' + str(i)].value) in lang_dict:
            ws_result['B' + str(i)].value = lang_dict[int(ws_result['A' + str(i)].value)]
            lang_dict.pop(int(ws_result['A' + str(i)].value))
            ws_result['A' + str(i)].fill = fil
            ws_result['B' + str(i)].fill = fil
        j += 1

    for k in lang_dict:
        ws_result['A' + str(j)].value = k
        ws_result['A' + str(j)].fill = fil
        ws_result['B' + str(j)].value = lang_dict[k]
        ws_result['B' + str(j)].fill = fil
        j += 1
    print(lang, '合并完成！\n')

print('正在保存修改...')
wb.save(path_tr)
print('文件处理完毕，请注意检查！')