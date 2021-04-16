import openpyxl

wb = openpyxl.load_workbook('Alert.xlsx')
sh = wb['design']

alert = []

for x in range(3,134):
    for y in range(3,31):
        print(x,y)
        if sh.cell(x,y).value > -1:
            alert.append((sh.cell(x,1).value,sh.cell(1,y).value,sh.cell(x,y).value,sh.cell(x,2).value))

sh2 = wb['Sheet2']
count = 1
for line in alert:
    sh2.cell(count + 4, 1).value = count
    if line[1] == 'EQUIP':
        sh2.cell(count + 4, 2).value = 'equip'
        sh2.cell(count + 4, 3).value = ''
    elif line[1] == 'HERO':
        sh2.cell(count + 4, 2).value = 'hero'
        sh2.cell(count + 4, 3).value = ''
    elif line[1] == 'PROP':
        sh2.cell(count + 4, 2).value = 'prop'
        sh2.cell(count + 4, 3).value = ''
    elif line[1] == 'BATTLE_PASS_1':
        sh2.cell(count + 4, 2).value = 'battle_pass_1'
        sh2.cell(count + 4, 3).value = ''
    elif line[1] == 'BATTLE_PASS_2':
        sh2.cell(count + 4, 2).value = 'battle_pass_2'
        sh2.cell(count + 4, 3).value = ''
    elif line[1] == 'EXP':
        sh2.cell(count + 4, 2).value = 'exp'
        sh2.cell(count + 4, 3).value = ''
    elif line[1] == 'VIP_EXP':
        sh2.cell(count + 4, 2).value = 'vip_exp'
        sh2.cell(count + 4, 3).value = ''
    else:
        sh2.cell(count + 4, 2).value = 'asset'
        sh2.cell(count + 4, 3).value = line[1]
    sh2.cell(count + 4, 4).value = line[0]
    sh2.cell(count + 4, 5).value = line[2]
    sh2.cell(count + 4, 6).value = line[3]
    count += 1
# print(alert)
wb.save('Alert.xlsx')