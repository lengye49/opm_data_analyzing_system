import openpyxl
import numpy as np

qualities = []
talents = []
academies = []
jobs = []
limiters = []
mechanics = []

for i in range(400):
    qualities.append([])
    for j in range(17):
        qualities[i].append(0)

    talents.append([])
    for j in range(41):
        talents[i].append(0)

    academies.append([])
    for j in range(351):
        academies[i].append(0)

    jobs.append([])
    for j in range(151):
        jobs[i].append(0)

    limiters.append([])
    for j in range(11):
        limiters[i].append(0)

    mechanics.append([])
    for j in range(31):
        mechanics[i].append(0)

wb = openpyxl.load_workbook('formation_hero.xlsx')
ws = wb['649']


def split_formation(formation_str):
    team_info = formation_str.split('$')

    summary = ''
    for member in team_info:

        hero_str = member.split('/')
        hero_pos = hero_str[0]
        hero_info = hero_str[1].split(';')
        is_mercenary = int(hero_info[0])
        if is_mercenary > 0:
            continue
        hero_id = int(hero_info[1])
        quality = int(hero_info[2])
        lv = int(hero_info[3])
        talent = int(hero_info[4])
        academy_job = hero_info[10].split(',')
        academy = int(academy_job[0])
        job = int(academy_job[1])
        limiter = int(hero_info[11])
        mechanical_core = int(hero_info[12])

        qualities[lv][quality] += 1
        talents[lv][talent] += 1
        academies[lv][academy] += 1
        jobs[lv][job] += 1
        limiters[lv][limiter] += 1
        mechanics[lv][mechanical_core] += 1


for col in ws['A']:
    if col.row == 1:
        continue
    print(col.row)
    split_formation(col.value)

wb.close()

wb = openpyxl.load_workbook('lv_status.xlsx')
ws = wb['Sheet1']


def get_median_index(_list):
    mark = np.sum(_list) / 2
    t = 0
    for i in range(len(_list)):
        t += _list[i]
        if t >= mark:
            return i


for i in range(1, 400):

    ws.cell(i + 1, 1).value = i
    ws.cell(i + 1, 2).value = get_median_index(qualities[i])
    ws.cell(i + 1, 3).value = get_median_index(talents[i])
    ws.cell(i + 1, 4).value = get_median_index(academies[i])
    ws.cell(i + 1, 5).value = get_median_index(jobs[i])
    ws.cell(i + 1, 6).value = get_median_index(limiters[i])
    ws.cell(i + 1, 7).value = get_median_index(mechanics[i])

wb.save('lv_status.xlsx')
