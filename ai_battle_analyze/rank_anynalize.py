import pandas as pd
from openpyxl import load_workbook
import os


def _excel_add_sheet(_df, _writer, sht_name):
    book = load_workbook(_writer.path)
    _writer.book = book
    _df.to_excel(excel_writer=_writer, sheet_name=sht_name, index=False, startrow=1, startcol=1)
    _writer.close()


def add_title():
    wb = load_workbook(path)
    for sh in wb:
        sh['b1'] = droid_group_names[int(sh.title)]
    wb.save(path)
    wb.close()


def save_result(_df, _g):
    if os.path.exists(path):
        excel_writer = pd.ExcelWriter(path, engine='openpyxl')
        _excel_add_sheet(_df, excel_writer, str(_g))
    else:
        _df.to_excel(path, sheet_name=str(_g), index=False, startrow=1, startcol=1)


def get_hero_id(s):
    bot_id = int(s.split('_')[0])
    if is_droid:
        return int(bot_id / 100000) % 100
    else:
        return bot_id % 1000


def get_hero_pos(s):
    return int(s.split('_')[1])


def get_max_pos(_x):
    _df = _x.sort_values(by='win_prob', ascending=True)
    return int(_df.iloc[-1, :][2] + 1)


def get_hero_names():
    _path = 'rank_data/' + ('droid/droid.xlsx' if is_droid else 'hero/hero.xlsx')
    _df = pd.read_excel(_path, header=0, index_col=0)
    return _df


def get_data(ver, group):
    _path = 'rank_data/'
    _path += 'droid/' if is_droid else 'hero/'
    _path += str(ver) + '/hero-postion-win-rate-' + str(group) + '.txt'
    _df = pd.read_table(_path, sep=',')
    return _df


# 获取精选站位的胜率
def get_pos_win_prob():
    pass


def get_hero_rank(_df):
    _df['hero'] = _df.apply(
        lambda col: get_hero_id(col['hero_id']), axis=1)
    _df['pos'] = _df.apply(
        lambda col: get_hero_pos(col['hero_id']), axis=1)
    # if is_pos:
    #     _df['win_prob'] = _df.loc[((_df['hero']==1) & (_df['pos']==2)),'win_prob']
    _df = _df.drop(['hero_id', 'win_count', 'loss_count', 'win_loss_ratio', 'battle_count'], axis=1)
    _df1 = _df.groupby('hero').agg({'win_prob': 'median'})
    # if is_pos:
    #     _df1['win_prob'] = _df.loc[
    #         (_df['hero'] == _df1['hero'] & _df['pos'] == df2.loc[_df1['hero'], 'max_rate_pos']), 'win_prob']
    _df1['max_rate_pos'] = _df.groupby('hero').apply(get_max_pos)
    return _df1


def handle_data(_df):
    _df = get_hero_rank(_df)
    _df = _df.sort_values(by='win_prob', ascending=False)
    _df = _df.reset_index()
    _df['rank'] = _df.index.get_level_values(0).values + 1
    _df = _df.set_index(_df['hero'], drop=True)
    return _df


labels = ['id', '英雄名称', '中位胜率', '最优站位', '本轮排名', '上轮排名', '排名变化']
command = input('请输入比对信息：\n格式：当前版本号路径,对比版本号路径,战斗类型(0英雄|1机器人)，例如26307_pos,26307,0\n')
droid_group_names = [
    '1星60级', '1星120级', '2星60级', '2星120级', '3星60级', '3星120级', '4星60级', '4星120级', '5星60级', '5星120级']
hero_group_names = ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']

_com = command.split(',')
if len(_com) != 3:
    print('输入格式有误')
else:
    is_droid = (_com[2] == '1')
    ver_new = _com[0]
    is_pos = ('pos' in ver_new)
    ver_old = _com[1]
    path = 'rank_data/rank_' + ver_new + '_vs_' + ver_old + '.xlsx'
    max_group = 10 if is_droid else 16
    df0 = get_hero_names()
    for _g in range(0, max_group):
        # 先计算老版本的，如果是精选站位可以使用
        df2 = handle_data(get_data(ver_old, _g))
        df1 = handle_data(get_data(ver_new, _g))
        # 将df1中有而df2中没有的英雄，用基础数据补足
        for x in list(df1.index):
            if x not in list(df2.index):
                df2 = df2.append(pd.DataFrame({'hero': x, 'win_prob': 0, 'max_rate_pos': 0, 'rank': 0},
                                 index=[x], copy=True), ignore_index=True)
        df1['old_rank'] = df2.loc[df1['hero'], 'rank']
        df1['rank_change'] = df1['old_rank'] - df1['rank']
        df1['hero_name'] = df0.loc[df1['hero'], 'name']
        df1 = df1[['hero', 'hero_name', 'win_prob', 'max_rate_pos', 'rank', 'old_rank', 'rank_change']]
        df1.columns = labels
        save_result(df1, _g)
    add_title()
